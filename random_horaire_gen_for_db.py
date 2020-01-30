from openpyxl import load_workbook, Workbook
import random

horaires_employes = {}
for i in range(1,11):
    wb = load_workbook('data/horaire{}.xlsx'.format(i))
    horaires_employes["{}".format(i)] = wb
dispo = ["J","S","OK","N"]
for employe in horaires_employes:
    wb = horaires_employes[employe]
    ws = wb.active
    print("book{}".format(employe))
    for row in ws.iter_rows(4,4,1,37):
        for cell in row:
            randomValue = random.sample(dispo,1)
            cell.value = randomValue[0]
    wb.save('data/horaires{}.xlsx'.format(employe))







